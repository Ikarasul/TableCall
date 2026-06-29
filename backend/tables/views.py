"""
tables/views.py
===============
API Views สำหรับ Tables app ของ TableCall

Endpoints:
- GET  /api/tables/                         → Floor View (ทุกโต๊ะ + pending count)
- GET  /api/tables/<qr_token>/              → Table detail
- POST /api/tables/<qr_token>/notify/       → ลูกค้าเรียกพนักงาน (with Redis throttle 30s)

Business Logic:
- Redis throttle: 1 request / 30 วินาที / (qr_token + kind)
- Key: notify_throttle:{qr_token}:{kind}
- สร้าง Notification แล้ว broadcast ผ่าน Channel Layer
"""

import logging
import threading
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.conf import settings
from django.core.cache import cache
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from utils.query_token_auth import QueryParamJWTAuthentication

from .models import RestaurantTable
from .serializers import TableSerializer, TableDetailSerializer, NotifyRequestSerializer
from notifications.models import Notification
from notifications.serializers import NotificationSerializer
from notifications.fcm import send_push_to_staff

logger = logging.getLogger('tablecall')


def _throttle_key(qr_token: str, kind: str) -> str:
    """
    Redis key สำหรับ throttle การเรียกพนักงาน
    format: notify_throttle:{qr_token}:{kind}
    """
    return f'notify_throttle:{qr_token}:{kind}'


# ---------------------------------------------------------------------------
# 1. GET /api/tables/ → Floor View
# ---------------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def table_list(request):
    """
    แสดงโต๊ะทั้งหมดพร้อม pending notification count
    ใช้ใน Staff Floor View (หน้าภาพรวมห้องอาหาร)
    
    Response: list of TableSerializer
    """
    tables = RestaurantTable.objects.filter(is_active=True).prefetch_related('notifications').order_by('sort_order', 'number')
    serializer = TableSerializer(tables, many=True)
    return Response(serializer.data)


# ---------------------------------------------------------------------------
# 2. GET /api/tables/<qr_token>/ → Table Detail
# ---------------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([AllowAny])
def table_detail(request, qr_token: str):
    """
    แสดงรายละเอียดโต๊ะจาก qr_token
    ลูกค้าใช้เพื่อดูว่าสแกน QR ถูกต้องและโต๊ะที่ตัวเองนั่ง
    
    Permission: AllowAny (ลูกค้าไม่ต้อง login)
    """
    try:
        table = RestaurantTable.objects.get(qr_token=qr_token, is_active=True)
    except RestaurantTable.DoesNotExist:
        return Response(
            {'detail': 'ไม่พบโต๊ะนี้ หรือ QR Code ไม่ถูกต้อง'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception:
        return Response(
            {'detail': 'QR Token ไม่ถูกต้อง'},
            status=status.HTTP_400_BAD_REQUEST
        )

    serializer = TableDetailSerializer(table)
    return Response(serializer.data)


# ---------------------------------------------------------------------------
# 3. POST /api/tables/<qr_token>/notify/ → ลูกค้าเรียกพนักงาน
# ---------------------------------------------------------------------------
@api_view(['POST'])
@permission_classes([AllowAny])
def table_notify(request, qr_token: str):
    """
    ลูกค้ากดปุ่มเรียกพนักงาน / เรียกเก็บเงิน
    
    Body: { "kind": "call" | "bill" }
    
    Flow:
    1. ตรวจสอบ qr_token ว่าโต๊ะมีอยู่จริง
    2. ตรวจสอบ Redis throttle (30 วินาที)
       - ถ้าอยู่ใน cooldown → 429
    3. สร้าง Notification record
    4. Set Redis throttle key (TTL = 30 วินาที)
    5. Broadcast ผ่าน Channel Layer ไปยัง staff group
    6. Return notification data
    
    Response 429: ถ้ากด too fast
    Response 201: notification สร้างสำเร็จ
    """
    # ตรวจสอบโต๊ะ
    try:
        table = RestaurantTable.objects.get(qr_token=qr_token, is_active=True)
    except RestaurantTable.DoesNotExist:
        return Response(
            {'detail': 'ไม่พบโต๊ะนี้ หรือ QR Code ไม่ถูกต้อง'},
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception:
        return Response(
            {'detail': 'QR Token ไม่ถูกต้อง'},
            status=status.HTTP_400_BAD_REQUEST
        )

    # Validate request body
    serializer = NotifyRequestSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    kind = serializer.validated_data['kind']
    qr_token_str = str(qr_token)

    # ตรวจสอบ throttle ใน Redis — ใช้ cache.add() เพื่อ atomic set-if-not-exists
    throttle_key = _throttle_key(qr_token_str, kind)
    # cache.add() คืน True ถ้า key ยังไม่มี (set สำเร็จ), False ถ้ามีแล้ว
    throttle_set = cache.add(throttle_key, True, timeout=settings.NOTIFY_THROTTLE_SECONDS)

    if not throttle_set:
        # ยังอยู่ใน cooldown window
        ttl_remaining = cache.ttl(throttle_key)
        # cache.ttl() อาจคืน None (backend ไม่รองรับ) หรือค่าติดลบ ให้ fallback ที่ค่า default
        if ttl_remaining is None or ttl_remaining < 0:
            ttl_remaining = settings.NOTIFY_THROTTLE_SECONDS
        kind_display = 'เรียกพนักงาน' if kind == 'call' else 'เรียกเก็บเงิน'
        logger.info(
            f'Table {table.number} throttled for kind={kind}. '
            f'TTL remaining: {ttl_remaining}s'
        )
        return Response(
            {
                'detail': f'กรุณารอ {ttl_remaining} วินาที ก่อน{kind_display}อีกครั้ง',
                'retry_after': ttl_remaining,
                'throttled': True,
            },
            status=status.HTTP_429_TOO_MANY_REQUESTS
        )

    # สร้าง Notification
    notification = Notification.objects.create(
        table=table,
        kind=kind,
        status='pending',
    )
    logger.info(
        f'Notification created: id={notification.id}, '
        f'table={table.number}, kind={kind}'
    )

    # Broadcast notification ไปยัง staff ผ่าน Channel Layer
    _broadcast_notification_created(notification)

    notification_data = NotificationSerializer(notification).data
    # เพิ่ม cooldown_until เพื่อให้ frontend แสดง countdown
    cooldown_seconds = settings.NOTIFY_THROTTLE_SECONDS
    from datetime import datetime
    cooldown_until = datetime.fromisoformat(notification.created_at.isoformat()).timestamp() + cooldown_seconds
    notification_data['cooldown_until'] = notification.created_at.isoformat()
    notification_data['cooldown_seconds'] = cooldown_seconds
    return Response(notification_data, status=status.HTTP_201_CREATED)


def _broadcast_notification_created(notification: Notification) -> None:
    """
    ส่งสัญญาณแจ้งเตือนพนักงานแบบ 2 ช่องทางคู่ขนาน:
    1. WebSocket (Django Channels) → แอปที่เปิดอยู่หน้าจอ
    2. FCM Push Notification       → แอปที่อยู่ background หรือปิดอยู่
    """
    # --- 1. WebSocket Broadcast ---
    channel_layer = get_channel_layer()
    if channel_layer is None:
        logger.warning('Channel layer ไม่พร้อมใช้งาน ไม่สามารถ broadcast ได้')
    else:
        try:
            async_to_sync(channel_layer.group_send)(
                'staff_notifications',
                {
                    'type': 'notification.created',
                    'data': {
                        'notification': NotificationSerializer(notification).data
                    },
                }
            )
            logger.debug(f'Broadcast notification.created via WebSocket: id={notification.id}')
        except Exception as e:
            # ไม่ให้ error ใน channel layer ทำให้ request ล้มเหลว
            logger.error(f'WebSocket broadcast error: {e}')

    # --- 2. FCM Push Notification (ทำใน background thread ไม่บล็อก response) ---
    def _send_fcm():
        send_push_to_staff(
            table_number=notification.table.number,
            table_id=notification.table.id,
            notification_id=notification.id,
            kind=notification.kind,
            timestamp=notification.created_at.isoformat(),
        )

    fcm_thread = threading.Thread(target=_send_fcm, daemon=True)
    fcm_thread.start()


# ---------------------------------------------------------------------------
# 4. Admin API Endpoints
# ---------------------------------------------------------------------------
from staff.permissions import IsAdminStaff
from .serializers import AdminTableSerializer

@api_view(['GET', 'POST'])
@permission_classes([IsAdminStaff])
def admin_list_create_tables(request):
    """
    GET: ดูรายการโต๊ะทั้งหมด (รวมที่ปิดใช้งาน)
    POST: สร้างโต๊ะใหม่
    """
    if request.method == 'GET':
        tables = RestaurantTable.objects.all().order_by('sort_order', 'number')
        serializer = AdminTableSerializer(tables, many=True)
        return Response(serializer.data)
        
    elif request.method == 'POST':
        serializer = AdminTableSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['PATCH', 'DELETE'])
@permission_classes([IsAdminStaff])
def admin_detail_table(request, table_id: int):
    """
    PATCH: แก้ไขโต๊ะ (เช่น เปลี่ยนชื่อโต๊ะ, จำนวนที่นั่ง, หรือ ปิด/เปิด การใช้งาน)
    DELETE: ลบโต๊ะ (ลบออกจากระบบ)
    """
    try:
        table = RestaurantTable.objects.get(id=table_id)
    except RestaurantTable.DoesNotExist:
        return Response({'detail': 'ไม่พบโต๊ะนี้'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'PATCH':
        serializer = AdminTableSerializer(table, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        
    elif request.method == 'DELETE':
        table.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


# ---------------------------------------------------------------------------
# 5. Feedback API Endpoints
# ---------------------------------------------------------------------------
import csv
from django.http import HttpResponse
from .models import CustomerFeedback
from .serializers import CustomerFeedbackSerializer

@api_view(['POST'])
@permission_classes([AllowAny])
def submit_feedback(request, qr_token: str):
    """
    ลูกค้าส่งข้อเสนอแนะและความพึงพอใจ
    """
    try:
        table = RestaurantTable.objects.get(qr_token=qr_token)
    except RestaurantTable.DoesNotExist:
        return Response({'detail': 'QR Token ไม่ถูกต้อง'}, status=status.HTTP_404_NOT_FOUND)
    
    serializer = CustomerFeedbackSerializer(data=request.data)
    if serializer.is_valid():
        serializer.save(table=table)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAdminStaff])
@authentication_classes([QueryParamJWTAuthentication, JWTAuthentication])
def export_feedback_xlsx(request):
    """
    Admin ดาวน์โหลดข้อมูลเป็นไฟล์ Excel (XLSX)
    - ถ้าส่ง staff_id: ดาวน์โหลดประวัติการรับเรื่องของพนักงานนั้น (Staff Performance)
    - ถ้าไม่ส่ง: ดาวน์โหลดแบบประเมินความพึงพอใจทั้งหมด (Customer Feedback)
    """
    import openpyxl
    from openpyxl.styles import Font

    staff_id = request.GET.get('staff_id')
    
    wb = openpyxl.Workbook()
    ws = wb.active

    if staff_id:
        from notifications.models import Notification
        from staff.models import Staff
        try:
            staff = Staff.objects.get(id=staff_id)
        except Staff.DoesNotExist:
            return Response({'detail': 'Staff not found'}, status=404)
            
        filename = f"staff_performance_{staff.name}.xlsx"
        ws.title = "Performance"
        
        # Header for Staff Performance
        headers = ['วัน-เวลาที่เรียก', 'โต๊ะ', 'ประเภทการเรียก', 'เวลารับเรื่อง', 'เวลาตอบสนอง (วินาที)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        notifications = Notification.objects.filter(handled_by=staff).select_related('table').order_by('-created_at')
        for n in notifications:
            table_num = n.table.number if n.table else 'ไม่ทราบ'
            created_at_local = n.created_at.strftime('%Y-%m-%d %H:%M:%S')
            handled_at_local = n.handled_at.strftime('%Y-%m-%d %H:%M:%S') if n.handled_at else '-'
            kind_display = 'เรียกพนักงาน' if n.kind == 'call' else 'เช็คบิล'
            ws.append([created_at_local, table_num, kind_display, handled_at_local, n.response_time_seconds or '-'])
    else:
        # Export all customer feedback (with optional date filter)
        date_str = request.GET.get('date')
        filename = f"customer_feedback_{date_str}.xlsx" if date_str else "customer_feedback.xlsx"
        ws.title = "Customer Feedback"
        
        headers = ['วัน-เวลา', 'โต๊ะ', 'คะแนนความพึงพอใจ (1-5)', 'ข้อเสนอแนะ']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        feedbacks = CustomerFeedback.objects.all().select_related('table').order_by('-created_at')
        
        if date_str:
            import datetime as dt
            try:
                date_obj = dt.datetime.strptime(date_str, '%Y-%m-%d').date()
                feedbacks = feedbacks.filter(created_at__date=date_obj)
            except ValueError:
                pass
        
        for fb in feedbacks:
            table_num = fb.table.number if fb.table else 'ไม่ทราบ'
            created_at_local = fb.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ws.append([created_at_local, table_num, fb.rating, fb.suggestions])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

@api_view(['POST'])
@permission_classes([IsAdminStaff])
def admin_reorder_tables(request):
    """
    Admin รีออเดอร์โต๊ะ
    Request body: [{'id': 1, 'sort_order': 0}, {'id': 2, 'sort_order': 1}]
    """
    data = request.data
    if not isinstance(data, list):
        return Response({'detail': 'Invalid data format, expected a list'}, status=400)
    
    # Simple bulk update
    from .models import RestaurantTable
    for item in data:
        if 'id' in item and 'sort_order' in item:
            RestaurantTable.objects.filter(id=item['id']).update(sort_order=item['sort_order'])
            
    return Response({'detail': 'Reordered successfully'})
